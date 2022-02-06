import os
import openpyxl
from openpyxl import Workbook

## Documentation
## https://openpyxl.readthedocs.io/en/stable/


def main():
    
    wb = Workbook()                     # create a workbook
    ws0 = wb.active                     # select active sheet (Sheet by default)
    ws0.title = "Summary"               # rename sheet
    ws1 = wb.create_sheet("Sheet1")     # insert at the end (default)
        
    for sheet in wb:                    # iterating through all the sheets
        print(sheet.title)
        
    d = ws0.cell(row=4, column=2, value=10) # write 10 in cell B4 and get cell reference
    ws0['b4'] = 20                          # write 20 in cell B4
    ws1['b4'] = ws0['b4'].value                   # write ws0 (Summary) cell b4 value in ws1 (Sheet1) cell b4
    
    wb.save('my_wb.xlsx')               # save workbook to a file on current working directory
    

    
if __name__ == '__main__':
    os.chdir('D:\\OneDrive - ITG Technologies\\Desktop')
    main()
